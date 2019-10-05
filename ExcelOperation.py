from builtins import chr

from openpyxl import Workbook,load_workbook
import time

class ExcelOperation:
    def __init__(self):
        pass

    def excelFileRead(self,file):
        #创建一个列表用于储存excel数据
        list = []
        listItem =[]
        #打开Excel文件
        print('正在打开文件...')
        wb = load_workbook(file)
        print('成功！')
        #打开Excel表
        print('正在打开表...')
        ws = wb.get_sheet_by_name('Template')
        print('成功！')
        #获得总行、列数
        rows = ws.max_row
        cols = ws.max_column
        print('表格共%s行，%s列'%(rows,cols))
        #获得所有单元格数据，以行为单位
        print('正在读取表中的所有数据...')
        for row in ws.rows :
            for cell in row :
                print('正在把',cell.value,'添加到行列表')
                #添加每行列表
                listItem.append(cell.value)
            #添加到总列表
            print('正在把',listItem,'添加到总列表\n')
            list.append(listItem)
            #清空行列表
            listItem = []
        return list

    def f(self):
        def formulawrite(data):
            '''
            #数据预处理，删除None
            for d in data :
                for cell in d :
                    if cell != None :
                        nowD .append(cell)
                nowData.append(nowD)
                nowD = []
            data = nowData
            '''
            # 定义列字符串
            l = eo.listchar()
            # 定义推移表数据列表
            nowD = []
            nowData = []
            # 定义父erp
            fErp = ''
            t = time.localtime(time.time())
            for d in data:
                # 当为标题时写入
                if d[0] == '层次':
                    for i in range(1, 32):
                        d.append('%d/%d/%d' % (t.tm_year, t.tm_mon, i))
                # 当为总成写入
                if d[0].endswith(r'总成'):
                    # 获得父ERP号
                    fErp = str(d[3])
                    # 把父的ERP号写入推移表数据列表
                    nowD.append(d[3])
                    nowData.append(nowD)
                    nowD = []

                    for i in range(1, 32):
                        d.append('=SUMIFS(组立推移表!%s$3:%s$10240,组立推移表!$A$3:$A$10240,"%s",组立推移表!$D$3:$D$10240,"计划")' % (
                        str(l[i + 5]), str(l[i + 5]), fErp))
                # 当以.开关的层次写入
                elif d[0].startswith(r'.'):
                    for i in range(1, 32):
                        d.append('=VLOOKUP("%s",$D:$AT,%d,0)*INDIRECT("J"&ROW())' % (fErp, i + 8))

            return data, nowData

    def writeExcel(self,excelPath,sheetName,data):
        wb = Workbook()
        #激活workbook
        #wb.active
        #创建表
        wb.create_sheet(sheetName)
        ws = wb.get_sheet_by_name(sheetName)

        #写入数据
        for d in data:
            ws.append(d)

        # 保存
        fIetm = time.localtime(time.time())
        from locale import str
        wb.save(excelPath + 'TP' + str(fIetm.tm_year) + str(fIetm.tm_mon) + str(fIetm.tm_mday) + '.xlsx')
        return '写入成功！'
    #生成列字符串
    def listchar(self):
        fieldList = []
        i = 65
        j = 65
        k = 65
        while i<91:
            fieldList.append(chr(i))
            i +=1
        while j<91:
                fieldList.append('A'+chr(j))
                j +=1
        while k<91:
                fieldList.append('B'+chr(k))
                k +=1
        return fieldList